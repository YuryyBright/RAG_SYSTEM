
from pathlib import Path
from .base_reader import BaseReader

class XlsxReader(BaseReader):
    """
    Reader for Microsoft Excel (.xlsx) files.

    This reader reads a .xlsx file and returns its content as a string,
    with each row separated by a newline character and each column
    separated by a comma and a space.

    Example:
        >>> reader = XlsxReader()
        >>> reader.read(Path('example.xlsx'))
        'column1, column2, column3\n1, 2, 3\n4, 5, 6\n'
    """

    def read(self, path: Path) -> str:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            result = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    result.append(", ".join(str(cell.value or "") for cell in row))
            return "\n".join(result)
        except Exception as e:
            return f"Error reading XLSX file: {str(e)}"

